# webapp/app.py
from __future__ import annotations

import os
import sys

# Ensure project root is importable when running `python webapp/app.py`.
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BASE_DIR not in sys.path:
    sys.path.append(BASE_DIR)

from telebot import types  # если хочешь ещё и клавиатуру
from tgbot.bot_setup import bot  # тот же bot, что и в handlers
import html

from datetime import date, datetime, timedelta
from typing import Any, Dict, Optional, List

from flask import Flask, render_template, request, jsonify
from psycopg2.extras import RealDictRow

from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from tgbot.services import (  # type: ignore
    get_or_create_user,
    get_pair_by_user,
    get_wishlist_for_owner,
    add_wishlist_item,
    set_pair_cloud_url,
    set_pair_start_date,
    get_partner_alias_for_user,
    set_partner_alias_for_user,
)
from tgbot.db import fetchone, execute, execute_returning_one  # type: ignore
from tgbot.config import BOT_USERNAME  # type: ignore


app = Flask(__name__, template_folder="templates", static_folder="static")


# ===== Вспомогательные классы/функции =====


class TGUserWrapper:
    """
    Минималистичный wrapper для get_or_create_user:
    ожидает tg_user.id / .username / .first_name / .last_name.
    """

    def __init__(self, data: Dict[str, Any]) -> None:
        self.id = data["id"]
        self.username = data.get("username")
        self.first_name = data.get("first_name")
        self.last_name = data.get("last_name")


def serialize_date(value) -> Optional[str]:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return None


def fmt_date_ddmmyyyy(value: Optional[date]) -> Optional[str]:
    if not isinstance(value, date):
        return None
    return value.strftime("%d.%m.%Y")


def serialize_wishlist_item(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": row["id"],
        "title": row["title"],
        "description": row.get("description"),
        "url": row.get("url"),
        "is_done": bool(row.get("is_done")),
        "priority": row.get("priority") or "medium",
        "created_at": serialize_date(row.get("created_at")),
    }


def serialize_note(row: Dict[str, Any], current_user_id: int) -> Dict[str, Any]:
    return {
        "id": row["id"],
        "text": row["text"],
        "author_user_id": row["author_user_id"],
        "is_mine": row["author_user_id"] == current_user_id,
        "created_at": serialize_date(row.get("created_at")),
    }


def compute_relationship_stats(start: date) -> Dict[str, Any]:
    """
    Логика очень близка к тому, что у тебя в start_date_flow:
    считаем дни вместе, годы/месяцы, прогресс до годовщины, "красивые даты", большой юбилей.
    """
    today = date.today()
    if start > today:
        return {
            "start_date_iso": serialize_date(start),
            "start_date_human": fmt_date_ddmmyyyy(start),
            "future": True,
        }

    days_together = (today - start).days

    # годы + месяцы (по годовщинам)
    years = today.year - start.year
    if (today.month, today.day) < (start.month, start.day):
        years -= 1
    if years < 0:
        years = 0

    last_year_anniv = date(start.year + years, start.month, start.day)

    months = (today.year - last_year_anniv.year) * 12 + (
        today.month - last_year_anniv.month
    )
    if today.day < start.day:
        months -= 1
    if months < 0:
        months = 0

    # до следующей годовщины
    next_anniv = date(start.year + years + 1, start.month, start.day)
    days_until_next = (next_anniv - today).days

    total_period_days = (next_anniv - last_year_anniv).days or 1
    done_days = (today - last_year_anniv).days
    done_days = max(0, min(done_days, total_period_days))
    ratio = done_days / total_period_days
    percent = int(ratio * 100)

    # "красивые" даты
    milestone_days = [
        100,
        200,
        300,
        400,
        500,
        600,
        700,
        800,
        900,
        1000,
        1500,
        2000,
        2500,
        3000,
    ]
    next_milestone = None
    milestone_date = None
    days_to_milestone = None

    for d in milestone_days:
        if d > days_together:
            next_milestone = d
            milestone_date = start + timedelta(days=d)
            days_to_milestone = d - days_together
            break

    # крупный юбилей по годам (каждые 5 лет)
    next_big_year = ((years // 5) + 1) * 5
    big_anniv_date = date(start.year + next_big_year, start.month, start.day)
    days_to_big = (big_anniv_date - today).days

    is_anniversary_today = (
        years > 0
        and today.month == start.month
        and today.day == start.day
    )

    return {
        "start_date_iso": serialize_date(start),
        "start_date_human": fmt_date_ddmmyyyy(start),
        "future": False,
        "days_together": days_together,
        "years": years,
        "months": months,
        "is_anniversary_today": is_anniversary_today,
        "days_until_next": days_until_next,
        "percent_to_next": percent,
        "next_milestone_days": next_milestone,
        "next_milestone_date": serialize_date(milestone_date) if milestone_date else None,
        "next_milestone_days_left": days_to_milestone,
        "next_big_year": next_big_year,
        "next_big_year_date": fmt_date_ddmmyyyy(big_anniv_date),
        "next_big_year_days_left": days_to_big,
    }


def get_current_user_and_pair(payload: Dict[str, Any]):
    """
    Общий helper: из JSON достаём user, создаём/находим его в БД и пару.
    """
    user_data = payload.get("user")
    if not user_data or "id" not in user_data:
        return None, None, jsonify({"ok": False, "error": "USER_REQUIRED"}), 400

    tg_user = TGUserWrapper(user_data)
    user_id = get_or_create_user(tg_user)
    pair = get_pair_by_user(user_id)

    return user_id, pair, None, None

def notify_partner_about_new_note(pair, user_id: int, text: str) -> None:
    if pair["creator_user_id"] == user_id:
        partner_user_id = pair["partner_user_id"]
    else:
        partner_user_id = pair["creator_user_id"]

    if not partner_user_id:
        return

    partner = fetchone(
        "SELECT telegram_id FROM users WHERE id = %s",
        (partner_user_id,),
    )
    if not partner or not partner.get("telegram_id"):
        return

    tg_id = partner["telegram_id"]
    preview = text[:100] + "…" if len(text) > 100 else text
    safe_text = html.escape(preview, quote=False)

    notif_text = (
        "📝 <b>Новая совместная заметка!</b>\n\n"
        f"{safe_text}"
    )

    try:
        bot.send_message(tg_id, notif_text, parse_mode="HTML")
    except Exception as e:
        print(f"Failed to send note notification: {e}")


def notify_partner_about_new_wish(pair, user_id: int, title: str) -> None:
    """
    Отправить партнёру уведомление о новом желании
    (логика очень похожа на ту, что в bot.handle_pending -> wishlist_add).
    """
    # определяем id партнёра
    if pair["creator_user_id"] == user_id:
        partner_user_id = pair["partner_user_id"]
    else:
        partner_user_id = pair["creator_user_id"]

    if not partner_user_id:
        return

    # достаём telegram_id партнёра
    partner = fetchone(
        "SELECT telegram_id, username, first_name FROM users WHERE id = %s",
        (partner_user_id,),
    )
    if not partner or not partner.get("telegram_id"):
        return

    tg_id = partner["telegram_id"]

    # кто добавил желание — из WebApp, у нас нет message.from_user,
    # но в payload прилетает user с username / first_name
    who = "Партнёр"
    # если хочешь, можно прокинуть username из payload и передать его сюда отдельным аргументом

    safe_title = html.escape(title, quote=False)
    safe_who = html.escape(who, quote=False)

    notif_text = (
        "🎁 <b>Новое желание в списке партнера!</b>\n\n"
        f"<b>{safe_who}</b> добавил(а): <b>{safe_title}</b>"
    )

    try:
        bot.send_message(tg_id, notif_text, reply_markup=None, parse_mode="HTML")
    except Exception as e:
        print(f"Failed to send wishlist notification from webapp: {e}")

from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def build_wishlist_xlsx(items: list[dict], sheet_name: str = "Wishlist") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limit

    PRIORITY_LABELS = {"high": "Очень хочу", "medium": "Хочу", "low": "Несрочно"}

    headers = ["Желание", "Приоритет", "Ссылка", "Создано"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    for r in items:
        created_at = r.get("created_at")
        priority = r.get("priority") or "medium"
        ws.append(
            [
                r.get("title") or "",
                PRIORITY_LABELS.get(priority, priority),
                r.get("url") or "",
                (created_at.strftime("%d.%m.%Y") if created_at else ""),
            ]
        )

    # простая авто-ширина
    for col in range(1, len(headers) + 1):
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    # telebot любит, когда у файла есть name
    bio.name = "wishlist.xlsx"
    return bio

from tgbot.db import fetchall  # если есть

def send_wishlist_to_bot(pair: dict, owner_user_id: int, receiver_user_id: int, title: str = "Список желаний") -> None:
    """
    owner_user_id  — чей список выгружаем (my list или partner list)
    receiver_user_id — кому в Telegram отправляем файл (обычно текущему юзеру)
    """
    receiver = fetchone(
        "SELECT telegram_id, first_name, username FROM users WHERE id = %s",
        (receiver_user_id,),
    )
    if not receiver or not receiver.get("telegram_id"):
        return

    tg_id = receiver["telegram_id"]

    items = fetchall(
        """
        SELECT title, url, created_at, priority
        FROM wishlist_items
        WHERE pair_id = %s AND owner_user_id = %s
        ORDER BY created_at DESC
        """,
        (pair["id"], owner_user_id),
    )

    xlsx = build_wishlist_xlsx(items, sheet_name=title)

    caption = f"📄 {title}\nВсего: {len(items)}"
    bot.send_document(tg_id, xlsx, caption=caption)
# ===== Маршруты =====


@app.route("/")
def index():
    return render_template("index.html")


@app.post("/api/init")
def api_init():
    """
    Инициализация состояния WebApp:
    - пользователь
    - пара (если есть)
    - мой список
    - список партнёра
    - ссылка на диск
    - данные по дате отношений.
    """
    data = request.json or {}

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code

    if not pair:
        # пара ещё не создана, но юзер в БД уже есть
        return jsonify(
            {
                "ok": True,
                "has_pair": False,
                "user_id": user_id,
            }
        )

    # определяем партнёра
    if pair["creator_user_id"] == user_id:
        partner_id = pair["partner_user_id"]
    else:
        partner_id = pair["creator_user_id"]

    # свои желания
    my_items_raw = get_wishlist_for_owner(pair["id"], user_id)
    my_items = [serialize_wishlist_item(i) for i in my_items_raw]

    # желания партнёра
    partner_items: List[Dict[str, Any]] = []
    partner_info: Optional[Dict[str, Any]] = None

    if partner_id:
        partner_items_raw = get_wishlist_for_owner(pair["id"], partner_id)
        partner_items = [serialize_wishlist_item(i) for i in partner_items_raw]

        partner_info = fetchone(
            "SELECT id, username, first_name FROM users WHERE id = %s",
            (partner_id,),
        )

    # инфо по дате
    start_stats = None
    if pair.get("start_date"):
        start_stats = compute_relationship_stats(pair["start_date"])

    # линк на диск
    cloud_url = pair.get("cloud_drive_url")
    partner_alias = get_partner_alias_for_user(pair, user_id)

    # совместные заметки
    notes_raw = fetchall(
        "SELECT id, author_user_id, text, created_at FROM notes WHERE pair_id = %s ORDER BY created_at DESC",
        (pair["id"],),
    ) or []
    notes = [serialize_note(n, user_id) for n in notes_raw]

    return jsonify(
        {
            "ok": True,
            "has_pair": True,
            "user_id": user_id,
            "pair": {
                "id": pair["id"],
                "start_date": serialize_date(pair.get("start_date")),
                "start_stats": start_stats,
                "cloud_url": cloud_url,
                "partner_alias": partner_alias,
            },
            "partner": {
                "id": partner_id,
                "username": partner_info["username"] if partner_info else None,
                "first_name": partner_info["first_name"] if partner_info else None,
            }
            if partner_id
            else None,
            "my_wishlist": my_items,
            "partner_wishlist": partner_items,
            "notes": notes,
        }
    )


@app.post("/api/wishlist/add")
def api_wishlist_add():
    data = request.json or {}
    title = (data.get("title") or "").strip()

    if not title:
        return jsonify({"ok": False, "error": "TITLE_REQUIRED"}), 400

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    item = add_wishlist_item(pair["id"], user_id, title)
    item_serialized = serialize_wishlist_item(item)

    # ← вот тут дергаем уведомление
    try:
        notify_partner_about_new_wish(pair, user_id, title)
    except Exception as e:
        # чтобы падением нотификации не ломать API
        print(f"Failed to notify partner about new wishlist item (webapp): {e}")

    return jsonify({"ok": True, "item": item_serialized})


@app.post("/api/wishlist/delete")
def api_wishlist_delete():
    """
    Удаление желания из своего списка.
    JSON: { "user": {...}, "item_id": 123 }
    """
    data = request.json or {}
    item_id = data.get("item_id")

    if not isinstance(item_id, int):
        return jsonify({"ok": False, "error": "ITEM_ID_REQUIRED"}), 400

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    # удаляем только свои желания
    execute(
        "DELETE FROM wishlist_items WHERE id = %s AND owner_user_id = %s",
        (item_id, user_id),
    )

    return jsonify({"ok": True})


@app.post("/api/wishlist/set_link")
def api_wishlist_set_link():
    """
    Привязать ссылку к своему желанию.
    JSON: { "user": {...}, "item_id": 123, "url": "https://..." }
    """
    data = request.json or {}
    item_id = data.get("item_id")
    url = (data.get("url") or "").strip()

    if not isinstance(item_id, int):
        return jsonify({"ok": False, "error": "ITEM_ID_REQUIRED"}), 400
    if not (url.startswith("http://") or url.startswith("https://")):
        return jsonify({"ok": False, "error": "INVALID_URL"}), 400

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    # обновляем только своё желание
    execute(
        "UPDATE wishlist_items SET url = %s WHERE id = %s AND owner_user_id = %s",
        (url, item_id, user_id),
    )

    return jsonify({"ok": True})

@app.post("/api/wishlist/send_to_bot")
def api_wishlist_send_to_bot():
    data = request.json or {}

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    # что отправляем: мой список или список партнёра
    target = (data.get("target") or "me").strip()  # "me" | "partner"

    if target == "partner":
        # чей список выгружаем — партнёра
        if pair["creator_user_id"] == user_id:
            owner_user_id = pair["partner_user_id"]
        else:
            owner_user_id = pair["creator_user_id"]
        title = "Список партнёра"
    else:
        # мой список
        owner_user_id = user_id
        title = "Мой список"

    if not owner_user_id:
        return jsonify({"ok": False, "error": "NO_PARTNER"}), 400

    try:
        # отправляем файл тому, кто нажал кнопку (user_id)
        send_wishlist_to_bot(pair, owner_user_id=owner_user_id, receiver_user_id=user_id, title=title)
    except Exception as e:
        print(f"Failed to send wishlist xlsx to bot: {e}")
        return jsonify({"ok": False, "error": "SEND_FAILED"}), 500

    return jsonify({"ok": True})

@app.post("/api/wishlist/toggle_done")
def api_wishlist_toggle_done():
    """
    Переключить флаг is_done для своего желания.
    JSON: { "user": {...}, "item_id": 123, "done": true }
    """
    data = request.json or {}
    item_id = data.get("item_id")
    done = data.get("done")

    if not isinstance(item_id, int):
        return jsonify({"ok": False, "error": "ITEM_ID_REQUIRED"}), 400
    if not isinstance(done, bool):
        return jsonify({"ok": False, "error": "DONE_BOOL_REQUIRED"}), 400

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    execute(
        "UPDATE wishlist_items SET is_done = %s WHERE id = %s AND owner_user_id = %s",
        (done, item_id, user_id),
    )

    return jsonify({"ok": True})


@app.post("/api/wishlist/set_priority")
def api_wishlist_set_priority():
    """
    Установить приоритет для своего желания.
    JSON: { "user": {...}, "item_id": 123, "priority": "high" }
    Допустимые значения priority: "high", "medium", "low"
    """
    data = request.json or {}
    item_id = data.get("item_id")
    priority = (data.get("priority") or "").strip()

    if not isinstance(item_id, int):
        return jsonify({"ok": False, "error": "ITEM_ID_REQUIRED"}), 400
    if priority not in ("high", "medium", "low"):
        return jsonify({"ok": False, "error": "INVALID_PRIORITY"}), 400

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    execute(
        "UPDATE wishlist_items SET priority = %s WHERE id = %s AND owner_user_id = %s",
        (priority, item_id, user_id),
    )

    return jsonify({"ok": True})


@app.post("/api/cloud/set")
def api_cloud_set():
    """
    Установить/обновить ссылку на общий диск.
    JSON: { "user": {...}, "url": "https://..." }
    """
    data = request.json or {}
    url = (data.get("url") or "").strip()

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    if url and not (url.startswith("http://") or url.startswith("https://")):
        return jsonify({"ok": False, "error": "INVALID_URL"}), 400

    # пустая строка = удалить ссылку
    set_pair_cloud_url(pair["id"], url or None)

    return jsonify({"ok": True})


@app.post("/api/partner_alias/set")
def api_partner_alias_set():
    """
    Установить/обновить отображаемое имя партнёра для текущего пользователя.
    JSON: { "user": {...}, "alias": "..." }
    """
    data = request.json or {}
    alias_raw = (data.get("alias") or "").strip()
    alias = alias_raw[:64] if alias_raw else None

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    set_partner_alias_for_user(pair, user_id, alias)
    return jsonify({"ok": True, "alias": alias})


@app.post("/api/startdate/set")
def api_startdate_set():
    """
    Установить/обновить дату начала отношений.
    JSON: { "user": {...}, "date_str": "ДД.ММ.ГГГГ" }
    """
    data = request.json or {}
    date_str = (data.get("date_str") or "").strip()

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    if not date_str:
        return jsonify({"ok": False, "error": "DATE_REQUIRED"}), 400

    import re

    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", date_str)
    if not m:
        return jsonify({"ok": False, "error": "BAD_FORMAT"}), 400

    day, month, year = map(int, m.groups())
    try:
        d = date(year, month, day)
    except ValueError:
        return jsonify({"ok": False, "error": "INVALID_DATE"}), 400

    set_pair_start_date(pair["id"], d)
    stats = compute_relationship_stats(d)

    return jsonify(
        {
            "ok": True,
            "start_date": serialize_date(d),
            "start_stats": stats,
        }
    )


@app.post("/api/pair/delete")
def api_pair_delete():
    """
    Удалить пару (как в боте).
    JSON: { "user": {...} }
    """
    data = request.json or {}

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    pair_id = pair["id"]

    # собираем телеграм-id обоих, чтобы можно было уведомить, если захочешь
    pair_full = fetchone(
        """
        SELECT p.id, p.creator_user_id, p.partner_user_id,
               u1.telegram_id AS t1,
               u2.telegram_id AS t2
        FROM pairs p
        JOIN users u1 ON u1.id = p.creator_user_id
        LEFT JOIN users u2 ON u2.id = p.partner_user_id
        WHERE p.id = %s
        """,
        (pair_id,),
    )

    execute("DELETE FROM pairs WHERE id = %s", (pair_id,))

    return jsonify({"ok": True})

@app.post("/api/wishlist/clear")
def api_wishlist_clear():
    """
    Очистить весь список желаний текущего пользователя в паре.
    JSON: { "user": {...} }
    """
    data = request.json or {}

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    # Вариант 1 (если в таблице есть колонка pair_id — наиболее корректно):
    execute(
        "DELETE FROM wishlist_items WHERE pair_id = %s AND owner_user_id = %s",
        (pair["id"], user_id),
    )

    # Вариант 2 (на 100% совместим с тем, что уже видно в коде — чистим все желания юзера):
    # execute(
    #     "DELETE FROM wishlist_items WHERE owner_user_id = %s",
    #     (user_id,),
    # )

    return jsonify({"ok": True})

@app.post("/api/notes/add")
def api_notes_add():
    data = request.json or {}
    text = (data.get("text") or "").strip()

    if not text:
        return jsonify({"ok": False, "error": "TEXT_REQUIRED"}), 400

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    note = execute_returning_one(
        "INSERT INTO notes (pair_id, author_user_id, text) VALUES (%s, %s, %s) RETURNING id, author_user_id, text, created_at",
        (pair["id"], user_id, text),
    )
    note_serialized = serialize_note(note, user_id)

    try:
        notify_partner_about_new_note(pair, user_id, text)
    except Exception as e:
        print(f"Failed to notify partner about new note: {e}")

    return jsonify({"ok": True, "note": note_serialized})


@app.post("/api/notes/delete")
def api_notes_delete():
    data = request.json or {}
    note_id = data.get("note_id")

    if not isinstance(note_id, int):
        return jsonify({"ok": False, "error": "NOTE_ID_REQUIRED"}), 400

    user_id, pair, err_resp, err_code = get_current_user_and_pair(data)
    if err_resp is not None:
        return err_resp, err_code
    if not pair:
        return jsonify({"ok": False, "error": "NO_PAIR"}), 400

    execute(
        "DELETE FROM notes WHERE id = %s AND author_user_id = %s",
        (note_id, user_id),
    )

    return jsonify({"ok": True})


if __name__ == "__main__":
    # dev-режим
    app.run(host="0.0.0.0", port=8000, debug=True)